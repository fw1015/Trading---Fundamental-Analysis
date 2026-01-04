import numpy as np
from openpyxl import load_workbook
# HISTORIC ASSET RETURNS
# [[EXPECTED RETURN],
# [ANNUALIZED RETURN],
# [ANNUALIZED STANDARD DEVIATION],
# [ANNUALIZED SHARPE RATIO (Rf = 0%)]]

def his_asset_returns(income_statement, cashflow_statement, enterprise_value, target, output, num_yr):
    year = min(len(income_statement), len(cashflow_statement), len(enterprise_value))
    re_list = [[]]
    ex_re_list = []
    an_re_list = []
    an_st_div_list = []
    an_sh_r_list = []

    if target == "S&P500":
        yr_re = s_n_p_returns(year)
    else:
        yr_re = target_re(income_statement, cashflow_statement, enterprise_value, year)

    yr_re_add = [x + 1 for x in yr_re]

    for yr in range(num_yr):
        ex_re = np.mean(yr_re[yr:])
        ex_re_list.append(ex_re)

        an_re = ((np.prod(yr_re_add[yr:])) ** (1 / (len(yr_re) - yr))) - 1
        an_re_list.append(an_re)

        an_st_div = np.std(yr_re[yr:])
        an_st_div_list.append(an_st_div)

        an_sh_r = an_re / an_st_div
        an_sh_r_list.append(an_sh_r)

    re_list.append(ex_re_list)
    re_list.append(an_re_list)
    re_list.append(an_st_div_list)
    re_list.append(an_sh_r_list)

    re_list.pop(0)

    if output == "normal":
        return re_list
    else:
        return yr_re

def s_n_p_returns(min_yr):
    s_n_p_re = []

    wb = load_workbook(filename="Market Historic Data.xlsx")
    wb.active
    sheets = wb.sheetnames
    s_n_p_ws = wb[sheets[0]]

    for i in range(6, 6 + min_yr):
        s_n_p_re.append(s_n_p_ws[f"I{i}"].value)

    return s_n_p_re

def target_re(income_statement, cashflow_statement, enterprise_value, min_yr):
    yr_re = []
    for yr in range(min_yr - 1):
        yr_inc = income_statement[yr]
        yr_csh = cashflow_statement[yr]
        yr_ev = enterprise_value[yr]
        pre_yr_ev = enterprise_value[yr + 1]

        date = yr_inc["calendarYear"]
        sh_out_dil = yr_inc["weightedAverageShsOutDil"]
        div_paid = yr_csh["dividendsPaid"] * -1
        sh_price = yr_ev["stockPrice"]
        pre_sh_price = pre_yr_ev["stockPrice"]

        div_per_sh = div_paid / sh_out_dil

        re = (sh_price + div_per_sh - pre_sh_price) / pre_sh_price

        yr_re.append(re)

    return yr_re

