from openpyxl import load_workbook

# YIELD RATIOS
# [[EARNINGS YIELD],
# [DIVIDENDS YIELD]
# [DIVIDENDS PAYOUT]]

def yield_ratios(income_statement, cashflow_statement, enterprise_value, num_yr):
    # year = len(income_statement)
    y_ratios = []

    ern_y_arr = []
    div_y_arr = []
    div_p_arr = []

    for yr in range(num_yr):
        yr_inc = income_statement[yr]
        yr_csh = cashflow_statement[yr]
        yr_ev = enterprise_value[yr]

        sh_price = yr_ev["stockPrice"]
        net_sales = yr_inc["netIncome"]
        sh_out_dil = yr_inc["weightedAverageShsOutDil"]

        ern_per_sh_dil = net_sales / sh_out_dil
        div_paid = - yr_csh["dividendsPaid"]
        div_per_sh_dil = div_paid / sh_out_dil

        # EARNINGS YIELD
        ern_y = ern_per_sh_dil / sh_price
        ern_y_arr.append(ern_y)

        # DIVIDENDS YIELD
        div_y = div_per_sh_dil / sh_price
        div_y_arr.append(div_y)

        # DIVIDENDS PAYOUT
        div_p = div_per_sh_dil / ern_per_sh_dil
        div_p_arr.append(div_p)

    y_ratios.append(ern_y_arr)
    y_ratios.append(div_y_arr)
    y_ratios.append(div_p_arr)

    return y_ratios

def market_yield(year):
    wb = load_workbook(filename="Market Historic Data.xlsx")
    wb.active
    sheets = wb.sheetnames
    s_n_p_ws = wb[sheets[0]]
    gov_note_ws = wb[sheets[1]]

    avg_snp_ern_yield = 0
    avg_snp_div_yield = 0
    avg_snp_div_pay = 0
    avg_gov_yield = 0

    for i in range(6, 6 + year):
        avg_snp_ern_yield += s_n_p_ws[f"F{i}"].value

        avg_snp_div_yield += s_n_p_ws[f"G{i}"].value

        avg_snp_div_pay += s_n_p_ws[f"H{i}"].value

        avg_gov_yield += gov_note_ws[f"C{i}"].value

    avg_snp_ern_yield = avg_snp_ern_yield / year
    avg_snp_div_yield = avg_snp_div_yield / year
    avg_snp_div_pay = avg_snp_div_pay / year
    avg_gov_yield = avg_gov_yield / year

    return [[avg_snp_ern_yield, avg_snp_div_yield, avg_snp_div_pay], [avg_gov_yield, avg_gov_yield]]